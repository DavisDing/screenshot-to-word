import pandas as pd
from openpyxl import load_workbook
import os
from tkinter import filedialog, messagebox

class ExcelHandler:
    def __init__(self, logger, input_dir, root):
        self.logger = logger
        self.input_dir = input_dir
        self.file_path = None
        self.df = None
        self.root = root
        self.version = "基础版"
        self.result_col_index = None

    def _show_error(self, msg: str):
        self.root.attributes('-topmost', True)
        messagebox.showerror("错误", msg, parent=self.root)
        self.root.attributes('-topmost', False)

    def _show_info(self, msg: str):
        self.root.attributes('-topmost', True)
        messagebox.showinfo("提示", msg, parent=self.root)
        self.root.attributes('-topmost', False)

    def select_excel_file(self) -> bool:
        try:
            files = [f for f in os.listdir(self.input_dir) if f.lower().endswith('.xlsx')]
            if not files:
                self._show_error("excel_input 目录下未找到 Excel 文件")
                self.logger.log("excel_input目录无Excel文件")
                return False

            if len(files) == 1:
                self.file_path = os.path.join(self.input_dir, files[0])
                self.logger.log(f"自动选中Excel文件：{files[0]}")
            else:
                self.root.attributes('-topmost', True)
                self.file_path = filedialog.askopenfilename(
                    title="请选择要执行的 Excel 文件",
                    initialdir=self.input_dir,
                    filetypes=[("Excel 文件", "*.xlsx")],
                    parent=self.root
                )
                self.root.attributes('-topmost', False)
                if self.file_path:
                    self.logger.log(f"用户选择Excel文件：{os.path.basename(self.file_path)}")
                else:
                    self.logger.log("用户取消选择Excel文件")
                    return False
            return True
        except Exception as e:
            self._show_error(f"选择Excel文件异常：{e}")
            self.logger.log(f"选择Excel文件异常：{e}")
            return False

    def load_excel(self) -> bool:
        try:
            self.df = pd.read_excel(self.file_path, header=0)
            self.logger.log(f"成功加载Excel文件：{os.path.basename(self.file_path)}")
        except Exception as e:
            self._show_error(f"读取Excel失败：{e}")
            self.logger.log(f"读取Excel失败: {e}")
            return False

        if {'测试名称', '验证点', '步骤名称', '步骤描述', '预期结果', '测试结果'}.issubset(set(self.df.columns)):
            self.version = "步骤版"
            self.logger.log("检测到步骤版Excel格式")
        else:
            self.version = "基础版"
            self.logger.log("检测到基础版Excel格式")

        if self.df.shape[1] < 3:
            self._show_error("Excel文件列数不足，至少需要3列（测试名称、验证点、测试结果）")
            self.logger.log("Excel列数不足3列")
            return False

        try:
            wb = load_workbook(self.file_path)
            ws = wb.active
            for i, cell in enumerate(ws[1], start=1):
                if cell.value == "测试结果":
                    self.result_col_index = i
                    self.logger.log(f"检测到“测试结果”列在第 {i} 列")
                    break
            if self.result_col_index is None:
                self._show_error("Excel中未找到“测试结果”列")
                self.logger.log("Excel中未找到“测试结果”列")
                return False
        except Exception as e:
            self._show_error(f"解析Excel列头失败：{e}")
            self.logger.log(f"解析Excel列头失败：{e}")
            return False

        return True

    def get_pending_cases(self):
        self.logger.log("开始获取未执行用例")
        for idx, row in self.df.iterrows():
            status = str(row['测试结果']).strip().lower() if pd.notna(row['测试结果']) else ''
            if status not in ['已执行', 'pass', 'passed']:
                filename = str(row['测试名称']).strip() if pd.notna(row['测试名称']) else ''
                checkpoint = str(row['验证点']).strip() if pd.notna(row['验证点']) else ''
                self.logger.log(f"待执行用例: 行={idx}, 用例名={filename}, 验证点={checkpoint}")
                yield idx, filename, checkpoint

    def mark_case_executed(self, index: int):
        # 内存中DataFrame更新
        self.df.at[index, '测试结果'] = "已执行"
        try:
            if self.result_col_index is None:
                self._show_error("内部错误：测试结果列索引未初始化")
                self.logger.log("错误：result_col_index 为 None，可能未正确加载 Excel")
                return
            excel_row = index + 2  # DataFrame 第一行为表头
            wb = load_workbook(self.file_path)
            ws = wb.active
            ws.cell(row=excel_row, column=self.result_col_index).value = "已执行"
            wb.save(self.file_path)
            self.logger.log(f"更新用例行 {excel_row} 状态为 已执行")
        except PermissionError:
            self._show_error("写入Excel失败，文件被占用或无权限。请关闭Excel后重试。")
            self.logger.log("写入Excel失败，权限不足或文件被占用。")
        except Exception as e:
            self._show_error(f"写入Excel失败：{e}")
            self.logger.log(f"写入Excel失败：{e}")

    def get_step_cases(self):
        from collections import defaultdict
        case_dict = defaultdict(list)
        for idx, row in self.df.iterrows():
            if pd.notna(row['测试名称']) or pd.notna(row['验证点']):
                current_test = str(row['测试名称']).strip()
                current_check = str(row['验证点']).strip()
            case_dict[(current_test, current_check)].append({
                "index": idx,
                "步骤名称": str(row['步骤名称']).strip() if pd.notna(row['步骤名称']) else "",
                "步骤描述": str(row['步骤描述']).strip() if pd.notna(row['步骤描述']) else "",
                "预期结果": str(row['预期结果']).strip() if pd.notna(row['预期结果']) else ""
            })
        return case_dict